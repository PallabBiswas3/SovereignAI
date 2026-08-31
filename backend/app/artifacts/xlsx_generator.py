from __future__ import annotations

from pathlib import Path
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class XlsxGenerator:
    def generate(self, output: Path, title: str, rows: list[dict[str, object]]) -> Path:
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = title[:31]
        if rows:
            headers = list(rows[0])
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="214761")
            for row in rows:
                sheet.append([self._cell_value(row.get(header)) for header in headers])
            sheet.auto_filter.ref = sheet.dimensions
            sheet.freeze_panes = "A2"
            for column in sheet.columns:
                sheet.column_dimensions[column[0].column_letter].width = min(45, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        workbook.save(output)
        return output

    @staticmethod
    def _cell_value(value: object) -> object:
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value, ensure_ascii=False, sort_keys=True)
        return value
