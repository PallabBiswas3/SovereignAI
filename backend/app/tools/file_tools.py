from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from app.tools.base import Tool, ToolResult, ToolRisk


ALLOWED_EXTENSIONS = {
    ".txt", ".md", ".pdf", ".docx", ".xlsx", ".csv", ".json",
    ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".gif",
}


class SafeWorkspace:
    def __init__(self, root: Path) -> None:
        self.root = root.resolve()
        self.root.mkdir(parents=True, exist_ok=True)

    def resolve(self, relative_path: str, *, must_exist: bool = False) -> Path:
        candidate = (self.root / relative_path).resolve()
        if candidate != self.root and self.root not in candidate.parents:
            raise ValueError("Path traversal outside the configured workspace is blocked")
        if must_exist and not candidate.exists():
            raise FileNotFoundError(relative_path)
        return candidate


class ListFilesTool(Tool):
    name = "list_files"
    description = "List supported files under the workspace"
    risk = ToolRisk.low
    input_schema = {"type": "object", "properties": {"path": {"type": "string"}}}

    def __init__(self, workspace: SafeWorkspace) -> None:
        self.workspace = workspace

    async def execute(self, args: dict[str, Any]) -> ToolResult:
        try:
            base = self.workspace.resolve(str(args.get("path", ".")), must_exist=True)
            files = [
                str(path.relative_to(self.workspace.root)).replace("\\", "/")
                for path in base.rglob("*")
                if path.is_file() and path.suffix.lower() in ALLOWED_EXTENSIONS
            ]
            return ToolResult(success=True, output=sorted(files))
        except (ValueError, FileNotFoundError, OSError) as exc:
            return ToolResult(success=False, error=str(exc))


class ReadFileTool(Tool):
    name = "read_file"
    description = "Read a supported local workspace file"
    risk = ToolRisk.low
    input_schema = {"type": "object", "properties": {"path": {"type": "string"}}, "required": ["path"]}

    def __init__(self, workspace: SafeWorkspace, max_chars: int = 200_000) -> None:
        self.workspace = workspace
        self.max_chars = max_chars

    async def execute(self, args: dict[str, Any]) -> ToolResult:
        try:
            path = self.workspace.resolve(str(args["path"]), must_exist=True)
            if path.suffix.lower() not in ALLOWED_EXTENSIONS:
                raise ValueError(f"Unsupported file type: {path.suffix}")
            text = extract_text(path)
            return ToolResult(success=True, output={"path": str(path.relative_to(self.workspace.root)), "text": text[: self.max_chars]})
        except (KeyError, ValueError, FileNotFoundError, OSError) as exc:
            return ToolResult(success=False, error=str(exc))


class WriteFileTool(Tool):
    name = "write_file"
    description = "Write UTF-8 text to an approved workspace file"
    risk = ToolRisk.medium
    input_schema = {"type": "object", "properties": {"path": {"type": "string"}, "content": {"type": "string"}}, "required": ["path", "content"]}

    def __init__(self, workspace: SafeWorkspace) -> None:
        self.workspace = workspace

    async def execute(self, args: dict[str, Any]) -> ToolResult:
        try:
            path = self.workspace.resolve(str(args["path"]))
            if path.suffix.lower() not in {".txt", ".md", ".json", ".csv", ".py"}:
                raise ValueError("Text writer does not permit this extension")
            path.parent.mkdir(parents=True, exist_ok=True)
            content = str(args.get("content", ""))
            path.write_text(content, encoding="utf-8")
            relative = str(path.relative_to(self.workspace.root)).replace("\\", "/")
            return ToolResult(success=True, output={"bytes": len(content.encode("utf-8"))}, generated_files=[relative])
        except (KeyError, ValueError, OSError) as exc:
            return ToolResult(success=False, error=str(exc))


class SearchFilesTool(Tool):
    name = "search_files"
    description = "Search readable workspace documents for literal text"
    risk = ToolRisk.low
    input_schema = {"type": "object", "properties": {"query": {"type": "string"}}, "required": ["query"]}

    def __init__(self, workspace: SafeWorkspace) -> None:
        self.workspace = workspace

    async def execute(self, args: dict[str, Any]) -> ToolResult:
        query = str(args.get("query", "")).strip().lower()
        if not query:
            return ToolResult(success=False, error="A non-empty query is required")
        matches: list[dict[str, Any]] = []
        for path in self.workspace.root.rglob("*"):
            if not path.is_file() or path.suffix.lower() not in ALLOWED_EXTENSIONS:
                continue
            try:
                text = extract_text(path)
            except Exception:
                continue
            for number, line in enumerate(text.splitlines(), start=1):
                if query in line.lower():
                    matches.append({
                        "file": str(path.relative_to(self.workspace.root)).replace("\\", "/"),
                        "line": number,
                        "text": line[:500],
                    })
                    if len(matches) >= 100:
                        return ToolResult(success=True, output=matches)
        return ToolResult(success=True, output=matches)


class FileMetadataTool(Tool):
    name = "get_file_metadata"
    description = "Return safe file metadata"
    risk = ToolRisk.low
    input_schema = {"type": "object", "properties": {"path": {"type": "string"}}, "required": ["path"]}

    def __init__(self, workspace: SafeWorkspace) -> None:
        self.workspace = workspace

    async def execute(self, args: dict[str, Any]) -> ToolResult:
        try:
            path = self.workspace.resolve(str(args["path"]), must_exist=True)
            stat = path.stat()
            return ToolResult(success=True, output={"name": path.name, "extension": path.suffix.lower(), "size": stat.st_size, "modified": stat.st_mtime})
        except (KeyError, ValueError, FileNotFoundError, OSError) as exc:
            return ToolResult(success=False, error=str(exc))


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md", ".csv", ".json"}:
        return path.read_text(encoding="utf-8", errors="replace")
    if suffix == ".pdf":
        from pypdf import PdfReader

        return "\n\n".join(
            f"[PAGE {index}]\n{page.extract_text() or ''}"
            for index, page in enumerate(PdfReader(str(path)).pages, start=1)
        )
    if suffix == ".docx":
        from docx import Document

        return "\n".join(paragraph.text for paragraph in Document(str(path)).paragraphs)
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"# Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                lines.append(",".join("" if value is None else str(value) for value in row))
        return "\n".join(lines)
    if suffix in {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".gif"}:
        return f"[Image file: {path.name}; use OCR or vision tool for content]"
    raise ValueError(f"Unsupported file type: {suffix}")
