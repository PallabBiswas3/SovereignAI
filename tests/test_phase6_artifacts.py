from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from pptx import Presentation
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.artifacts.pptx_generator import PptxGenerator
from app.artifacts.xlsx_generator import XlsxGenerator
from app.core.database import Base
from app.rag.embeddings import LocalHashEmbeddingProvider
from app.rag.ingestion import KnowledgeIngestionService
from app.rag.retrieval import LocalRetriever
from app.workflows.inspection import InspectionWorkflow


def test_all_artifact_formats_and_flagship_workflow(tmp_path: Path) -> None:
    sop = tmp_path / "Maintenance_SOP.md"
    sop.write_text(
        "[PAGE 37]\n## Section 7.4\nVibration shall not exceed 6.0 mm/s RMS. Readings above 9.0 mm/s require the pump to be removed.\n"
        "[PAGE 38]\n## Section 7.5\nNormal bearing temperature is up to 80 C. Above 90 C requires shutdown.\n"
        "[PAGE 39]\n## Section 7.6\nNormal discharge pressure is 4.8 to 5.5 bar.", encoding="utf-8"
    )
    report = tmp_path / "Pump_Inspection_Report.md"
    report.write_text("Pump-102\nVibration: 8.2 mm/s\nBearing temperature: 86 C\nDischarge pressure: 4.4 bar", encoding="utf-8")
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        embeddings = LocalHashEmbeddingProvider()
        KnowledgeIngestionService(session, embeddings).ingest(sop)
        docx_path = tmp_path / "Approval_Note.docx"
        analysis = InspectionWorkflow(LocalRetriever(session, embeddings)).analyze(report, docx_path)

    xlsx_path = XlsxGenerator().generate(tmp_path / "findings.xlsx", "Findings", analysis.findings)
    pptx_path = PptxGenerator().generate(tmp_path / "briefing.pptx", "Pump-102", [{"title": "Disposition", "bullets": [analysis.recommendation]}])

    assert "planned maintenance" in analysis.recommendation.lower()
    assert analysis.sources[0]["section"] == "7.4"
    assert len(Document(docx_path).paragraphs) > 5
    assert load_workbook(xlsx_path).active.max_row == 4
    assert len(Presentation(pptx_path).slides) == 2
