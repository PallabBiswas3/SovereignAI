import asyncio
from pathlib import Path

from docx import Document
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from pptx import Presentation

from app.documents.evidence import MultiFileEvidenceProcessor
from app.main import app


def test_type_aware_multifile_evidence_processing(tmp_path: Path) -> None:
    note = tmp_path / "note.md"
    note.write_text("Pump inspection supporting note", encoding="utf-8")
    csv_path = tmp_path / "readings.csv"
    csv_path.write_text("tag,value\nVIB,8.2\nTEMP,\n", encoding="utf-8")
    xlsx_path = tmp_path / "register.xlsx"
    workbook = Workbook()
    workbook.active.append(["equipment", "owner"])
    workbook.active.append(["Pump-102", "Maintenance"])
    workbook.save(xlsx_path)

    records = asyncio.run(MultiFileEvidenceProcessor().process([note, csv_path, xlsx_path], "review"))
    assert [item.processor for item in records] == ["md-extractor", "csv-structured-profiler", "openpyxl-structured"]
    assert records[1].metadata["missing"]["value"] == 1
    assert all(item.provenance["file"] == item.file for item in records)


def test_management_package_contains_valid_docx_xlsx_and_pptx() -> None:
    with TestClient(app) as client:
        response = client.post("/api/tasks", json={
            "request": "Analyze this inspection against the maintenance SOP and create a management package with DOCX XLSX PPTX",
            "attachments": ["uploads/Pump_Inspection_Report.md"],
            "use_case": "engineering",
        })
        assert response.status_code == 200, response.text
        run = response.json()
        payloads = {item["name"]: client.get(item["url"]).content for item in run["artifacts"]}

    assert set(payloads) == {"approval_note.docx", "inspection_analysis.xlsx", "management_briefing.pptx"}
    package_root = Path("workspace/artifacts") / run["id"]
    assert Document(package_root / "approval_note.docx").paragraphs
    assert load_workbook(package_root / "inspection_analysis.xlsx").active.max_row >= 2
    assert len(Presentation(package_root / "management_briefing.pptx").slides) >= 3
    assert len(run["evidence_records"]) >= 2
